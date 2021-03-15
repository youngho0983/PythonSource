from openpyxl import Workbook
import openpyxl.utils.cell as utils
from openpyxl.styles import Alignment, Font

excel_file = Workbook()

print(excel_file.sheetnames)

excel_file.remove(excel_file["Sheet"])

excel_file.remove(excel_file["Sheet"])

sheet1 = excel_file.create_sheet(index=0, title="Column")

sheet2 = excel_file.create_sheet(index=0, title="매출표")

print(excel_file.sheetnames)

for col in sheet1.iter_cols(min_col=1, max_col=6, min_row=1, max_row=3):
    for each_cell in col:
        each_cell.value = utils.get_column_letter(each_cell.column)
        each_cell.alignment = Alignment(horizontal="right", vertical="center")

        each_cell.font = Font(
            bold=True, name="Arial", size=12, underline="single", color="1bb638"
        )

for col in sheet2.iter_cols(min_col=1, max_col=6, min_row=1, max_row=3):

    for each_cell in col:
        each_cell.value = utils.get_column_letter(each_cell)

        each_cell.alignment = Alignment(horizontal="right", vertical="center")

        each_cell.font = Font(italic=True, name="Consoras", size=10, color="ff0000")

excel_file.save("./crawl/data/test2.xlsx")